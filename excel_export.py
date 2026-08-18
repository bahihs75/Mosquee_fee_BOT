from __future__ import annotations

from datetime import date, datetime
from io import BytesIO
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


HEADERS = [
    "رقم الطلب",
    "اسم المستخدم المسجل",
    "رابط ملف تيليغرام",
    "المعرف الرقمي",
    "اسم المسجد / المهمة",
    "الولاية",
    "بداية المهمة",
    "نهاية المهمة",
    "مدة المهمة",
    "المبلغ المطلوب",
    "المبلغ المدفوع",
    "العملة",
    "طريقة الدفع",
    "ملاحظات الطلب",
    "ملاحظات الدفع",
    "من أكد الدفع",
    "تاريخ التأكيد",
    "الإصدار المعتمد",
]


def _date(value: Any) -> str:
    if not value:
        return ""
    if isinstance(value, (datetime, date)):
        return value.strftime("%Y-%m-%d %H:%M:%S") if isinstance(value, datetime) else value.isoformat()
    return str(value)


def build_confirmed_workbook(rows: list[dict[str, Any]]) -> BytesIO:
    wb = Workbook()
    ws = wb.active
    ws.title = "الطلبات المؤكدة"
    ws.sheet_view.rightToLeft = True
    ws.freeze_panes = "A2"
    ws.append(HEADERS)

    header_fill = PatternFill("solid", fgColor="1F4E78")
    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for row in rows:
        ws.append([
            row.get("public_id", ""),
            row.get("full_name", ""),
            row.get("profile_link", ""),
            row.get("user_id", ""),
            row.get("mosque_name", ""),
            row.get("wilaya", ""),
            _date(row.get("mission_start_date")),
            _date(row.get("mission_end_date")),
            row.get("duration_text", ""),
            float(row.get("amount_requested", 0)),
            float(row.get("paid_amount", 0)),
            row.get("currency", "DZD"),
            row.get("payment_method", ""),
            row.get("additional_details", ""),
            row.get("payment_note", ""),
            row.get("cashier_user_id", ""),
            _date(row.get("paid_at")),
            row.get("approved_version_no", row.get("version_no", 1)),
        ])

    for cell in ws[1]:
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
    for col in range(1, len(HEADERS) + 1):
        width = min(max(len(HEADERS[col - 1]) + 2, 14), 28)
        ws.column_dimensions[get_column_letter(col)].width = width
    ws.auto_filter.ref = ws.dimensions
    ws.row_dimensions[1].height = 34

    meta = wb.create_sheet("تعريف الحقول")
    meta.sheet_view.rightToLeft = True
    meta.append(["الحقل", "الشرح"])
    for cell in meta[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = header_fill
    definitions = [
        ("رقم الطلب", "المعرف الدائم للطلب في قاعدة البيانات."),
        ("اسم المستخدم المسجل", "الاسم الرسمي المحفوظ ولا يستطيع المستخدم تغييره بنفسه."),
        ("رابط ملف تيليغرام", "رابط عام إن وُجد username، أو رابط tg://user للمستخدم الخاص."),
        ("المبلغ المدفوع", "المبلغ الذي أكد الصندوق دفعه فعلياً."),
        ("الإصدار المعتمد", "نسخة الطلب التي اعتمدتها الإدارة قبل تحويله للصندوق."),
    ]
    for item in definitions:
        meta.append(item)
    meta.column_dimensions["A"].width = 28
    meta.column_dimensions["B"].width = 80
    for row in meta.iter_rows():
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output
