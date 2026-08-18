from decimal import Decimal

from openpyxl import load_workbook

from excel_export import build_confirmed_workbook


def test_confirmed_workbook_contains_financial_columns():
    stream = build_confirmed_workbook([
        {
            "public_id": "EXP-2026-000001",
            "full_name": "مستخدم تجريبي",
            "profile_link": "tg://user?id=123",
            "user_id": 123,
            "mosque_name": "مسجد الاختبار",
            "wilaya": "الجزائر",
            "duration_text": "3 أيام",
            "amount_requested": Decimal("100.00"),
            "paid_amount": Decimal("100.00"),
            "currency": "DZD",
            "payment_method": "نقداً",
            "additional_details": "",
            "payment_note": "تم التسليم",
            "cashier_user_id": 456,
            "approved_version_no": 1,
        }
    ])
    wb = load_workbook(stream)
    assert "الطلبات المؤكدة" in wb.sheetnames
    headers = [cell.value for cell in wb["الطلبات المؤكدة"][1]]
    assert "رابط ملف تيليغرام" in headers
    assert "المبلغ المدفوع" in headers
    assert wb["الطلبات المؤكدة"]["A2"].value == "EXP-2026-000001"
